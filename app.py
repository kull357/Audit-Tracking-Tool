
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from pytz import timezone
IST = timezone('Asia/Kolkata')
from datetime import datetime
import os
import re

app = Flask(__name__)
CORS(app)

EXCEL_FILE = 'data.xlsx'

# Initialize Excel with headers if not present
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    headers = [
        "ticket_id", "ticket_name", "BLI", "request_assigned_date", "request_due_date",
        "developer_name", "urls_list", "no_of_urls", "audit_type",
        "auditor_name", "audit_assigned_date", "audit_assigned_time",
        "audit_comments", "audit_status", "audit_completion_date", "audit_completion_time"
    ]
    ws.append(headers)
    wb.save(EXCEL_FILE)




@app.route('/submit_ticket', methods=['POST'])
def submit_ticket():
    data = request.json

    # 🛡️ Step 1: Ensure Excel file exists and has headers
    if not os.path.exists(EXCEL_FILE) or os.path.getsize(EXCEL_FILE) == 0:
        wb = Workbook()
        ws = wb.active
        headers = [
            "ticket_id", "ticket_name", "BLI", "request_assigned_date", "request_due_date",
            "developer_name", "urls_list", "no_of_urls", "audit_type",
            "auditor_name", "audit_assigned_date", "audit_assigned_time",
            "audit_comments", "audit_status", "audit_completion_date", "audit_completion_time"
        ]
        ws.append(headers)
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    ticket_id = data.get("ticket_id")
    now = datetime.now(IST)
    audit_date = now.strftime("%Y-%m-%d")
    audit_time = now.strftime("%H:%M:%S")

    # ✅ Prevent duplicate ticket_id with same time
    for row in ws.iter_rows(min_row=2, values_only=False):
        existing_id = row[0].value
        existing_time = row[11].value  # audit_assigned_time is at index 11
        if existing_id == ticket_id and existing_time == audit_time:
            return jsonify({"message": "Ticket ID already exists with same audit time"}), 400

    # ✅ Format URLs
    raw_urls = data.get("urls_list", "")
    normalized = re.sub(r'(https?://)', r' \1', raw_urls).strip()
    url_parts = re.split(r'[,\s]+', normalized)
    valid_urls = [url for url in url_parts if re.match(r'^https?://', url)]
    formatted_urls = ", ".join(valid_urls)

    # ✅ Append new row with correct column order
    new_row = [
        data.get("ticket_id"),               # 0
        data.get("ticket_name"),             # 1
        data.get("BLI"),                     # 2
        data.get("request_assigned_date"),   # 3
        data.get("request_due_date"),        # 4
        data.get("developer_name"),          # 5
        formatted_urls,                      # 6
        data.get("no_of_urls"),              # 7
        data.get("audit_type"),              # 8
        data.get("auditor_name"),            # 9 ✅ Fixed
        audit_date,                          # 10 ✅
        audit_time,                          # 11 ✅
        "",                                  # 12: audit_comments
        "",                                  # 13: audit_status
        "",                                  # 14: audit_completion_date
        "",                                  # 15: audit_completion_time
    ]

    ws.append(new_row)
    wb.save(EXCEL_FILE)
    return jsonify({"message": "Ticket submitted successfully"})










@app.route('/search_ticket', methods=['GET'])
def search_ticket():
    ticket_id = request.args.get('ticket_id')
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    results = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if str(row[0]) == ticket_id:
            if any(not row[j] for j in range(12, 16)):
                record = dict(zip(headers, row))
                record["Row Number"] = i
                results.append(record)

    if results:
        return jsonify(results)
    else:
        return jsonify({"message": "No incomplete records found for Ticket ID"}), 404

@app.route('/search_ticket_all', methods=['GET'])
def search_ticket_all():
    ticket_id = request.args.get('ticket_id')
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    all_rows = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if str(row[0]) == ticket_id:
            record = dict(zip(headers, row))
            record["Row Number"] = i
            all_rows.append(record)

    if all_rows:
        return jsonify(all_rows)
    else:
        return jsonify({"message": "No records found for Ticket ID"}), 404

@app.route('/search_row', methods=['GET'])
def search_row():
    row_number = int(request.args.get('row_number'))
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    if 2 <= row_number <= ws.max_row:
        row = list(ws.iter_rows(min_row=row_number, max_row=row_number, values_only=True))[0]
        headers = [cell.value for cell in ws[1]]
        data = dict(zip(headers, row))
        data["Row Number"] = row_number
        return jsonify(data)
    else:
        return jsonify({"message": "Row not found"}), 404

@app.route('/add_audit', methods=['POST'])
def add_audit():
    data = request.json
    row_number = int(data.get("row_number"))
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    if 2 <= row_number <= ws.max_row:
        row = list(ws.iter_rows(min_row=row_number, max_row=row_number))[0]
        # row[9].value = data.get("auditor_name")  <-- auditor_name is now handled in ticket form
        row[12].value = data.get("audit_comments")
        row[13].value = data.get("audit_status")
        now = datetime.now(IST)
        row[14].value = now.strftime("%Y-%m-%d")
        row[15].value = now.strftime("%H:%M:%S")
        wb.save(EXCEL_FILE)
        return jsonify({"message": "Audit data updated successfully for row number " + str(row_number)})
    else:
        return jsonify({"message": "Invalid row number"}), 404

@app.route('/download_report', methods=['GET'])
def download_report():
    if os.path.exists(EXCEL_FILE):
        return send_file(EXCEL_FILE, as_attachment=True)
    else:
        return jsonify({"message": "Excel file not found."}), 404
    

@app.route('/')
def home():
    return '✅ Audit Tracking Tool is Live! Use /submit_ticket or other endpoints.'


import os

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))  # Use the port Render gives, default to 5000
    app.run(host='0.0.0.0', port=port)
